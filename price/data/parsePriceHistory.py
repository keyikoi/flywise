#!/usr/bin/env python3
"""
航班价格历史数据导出脚本

从 Excel 文件中提取航班数据，计算每日平均价格，输出为 JSON 格式
"""

import openpyxl
import json
import glob
from datetime import datetime, timedelta

def find_excel_file():
    """查找 Excel 文件"""
    files = glob.glob("/Users/keyi/Desktop/*春运*.xlsx")
    for f in files:
        # 跳过临时文件
        if f.startswith("~$") or "/~$" in f:
            continue
        # 检查文件大小（临时文件通常很小）
        import os
        if os.path.getsize(f) < 1000:
            continue
        return f
    return None

def parse_cabin_data(cabin_json):
    """
    解析舱等数据，计算平均价格

    Args:
        cabin_json: 舱等数据 JSON 字符串或字典

    Returns:
        dict: {total_income, total_sales, avg_price}
    """
    if isinstance(cabin_json, str):
        try:
            data = json.loads(cabin_json)
        except:
            return None
    else:
        data = cabin_json

    if not data or 'BaseCabins' not in data:
        return None

    total_income = 0
    total_sales = 0

    for cabin in data['BaseCabins']:
        income = float(cabin.get('Income', 0) or 0)
        # 销售数 = BK 字段（订座数）
        sales = int(cabin.get('BK', 0) or 0)

        total_income += income
        total_sales += sales

    # 计算平均价格
    avg_price = round(total_income / total_sales) if total_sales > 0 else 0

    return {
        'total_income': total_income,
        'total_sales': total_sales,
        'avg_price': avg_price,
    }

def convert_to_int(value):
    """转换为整数，处理字符串和浮点数"""
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(float(value))
        except:
            return 0
    return 0

def convert_to_float(value):
    """转换为浮点数"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except:
            return 0.0
    return 0.0

def format_date(date_obj):
    """格式化日期为 YYYY-MM-DD"""
    if date_obj is None:
        return None
    if isinstance(date_obj, str):
        # 尝试解析字符串
        try:
            if ' ' in date_obj:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d %H:%M:%S')
            else:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
        except:
            return None
    return date_obj.strftime('%Y-%m-%d')

def main():
    print("=== 航班价格历史数据导出 ===\n")

    # 查找 Excel 文件
    excel_path = find_excel_file()
    if not excel_path:
        print("未找到 Excel 文件")
        return

    print(f"使用文件：{excel_path}")

    # 加载 Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 获取表头
    headers = [cell.value for cell in ws[1]]
    print(f"共 {len(headers)} 列")

    # 建立列索引映射
    col_idx = {h: i for i, h in enumerate(headers)}
    print(f"列索引：{col_idx}")

    # 处理数据
    price_data = []
    route_stats = {}  # 航线统计

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 提取关键字段
            flight_no = row[col_idx.get('航班号', 0)]
            flight_date = row[col_idx.get('航班日期', 1)]
            origin = row[col_idx.get('出发', 2)]
            destination = row[col_idx.get('到达', 3)]
            sales = convert_to_int(row[col_idx.get('销售数', 10)])
            passengers = convert_to_int(row[col_idx.get('上客人数', 9)])
            cabin_data = row[col_idx.get('舱等数据', 18)]

            # 跳过无效数据
            if not origin or not destination:
                continue

            # 解析舱等数据
            cabin_info = parse_cabin_data(cabin_data)
            if not cabin_info or cabin_info['avg_price'] == 0:
                continue

            # 格式化日期
            flight_date_str = format_date(flight_date)
            if not flight_date_str:
                continue

            # 创建记录
            record = {
                'flightNo': flight_no,
                'flightDate': flight_date_str,
                'origin': origin,
                'destination': destination,
                'route': f"{origin}-{destination}",
                'sales': sales,
                'passengers': passengers,
                'totalIncome': cabin_info['total_income'],
                'avgPrice': cabin_info['avg_price'],
            }

            price_data.append(record)

            # 统计航线
            route = f"{origin}-{destination}"
            if route not in route_stats:
                route_stats[route] = {'count': 0, 'dates': set()}
            route_stats[route]['count'] += 1
            route_stats[route]['dates'].add(flight_date_str)

        except Exception as e:
            print(f"行 {row_idx} 解析失败：{e}")
            continue

    # 输出统计
    print(f"\n=== 处理结果 ===")
    print(f"有效记录：{len(price_data)} 条")
    print(f"覆盖航线：{len(route_stats)} 条")

    # 显示航线统计
    print("\n=== 航线统计 ===")
    for route, stats in sorted(route_stats.items()):
        date_range = get_date_range(stats['dates'])
        print(f"  {route}: {stats['count']} 条记录, 日期范围：{date_range}")

    # 保存 JSON
    output_path = "/Users/keyi/work/claude-code-test/price/data/price_history.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(price_data, f, ensure_ascii=False, indent=2)

    print(f"\n已保存到：{output_path}")

    # 显示样本数据
    print("\n=== 样本数据 (前 5 条) ===")
    for record in price_data[:5]:
        print(f"  {record['flightDate']} {record['route']}: ¥{record['avgPrice']} (销售:{record['sales']}, 收入:¥{record['totalIncome']})")

def get_date_range(dates):
    """获取日期范围"""
    if not dates:
        return "无"
    sorted_dates = sorted(dates)
    return f"{sorted_dates[0]} ~ {sorted_dates[-1]}"

if __name__ == '__main__':
    main()
